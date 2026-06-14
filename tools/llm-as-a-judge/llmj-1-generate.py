#!/usr/bin/env python3

import os
import sys

import llmj

try:
    import openpyxl
except ImportError:
    llmj.abort_missing_package('openpyxl')

def process_prompt(in_runtime, in_path):
    try:
        with open(in_path, encoding='utf-8') as f:
            template = f.read()
    except Exception:
        print(f'ERROR : can not read "{in_path.name}"')
        llmj.abort()
    filename = in_path.name.removesuffix(llmj.SUFFIX_PRO) + llmj.SUFFIX_GEN
    dst_path = in_path.with_name(filename)
    if dst_path.exists():
        workbook = openpyxl.load_workbook(dst_path)
    else:
        workbook = openpyxl.Workbook()
    sheet = workbook.active
    colDict = {}
    for key in llmj.TERM_GEN:
        colDict[key] = llmj.find_append_column(sheet, llmj.TERM_GEN[key])
    sourceArr = sorted(llmj.DIR_SOURCE.glob('*.txt'))
    for row, src_path in enumerate(sourceArr, start=2):
        print(f'processing  : {row - 1} / {len(sourceArr)}')
        if sheet.cell(row, colDict['GENERATED']).value:
            continue
        textDict = {}
        try:
            with open(src_path, encoding='utf-8') as f:
                textDict['ORIGINAL'] = f.read()
        except Exception:
            print(f'ERROR : can not read "{src_path.name}"')
            llmj.abort()
        prompt = template.replace(llmj.ORIGINAL_PLACEHOLDER, textDict['ORIGINAL'])
        textDict['GENERATED'] = llmj.invoke_llm(in_runtime, llmj.LLM_MODEL, prompt)
        for key in llmj.TERM_GEN:
            sheet.cell(row, colDict[key]).value = textDict[key]
        workbook.save(dst_path)
    print(f'generated : {dst_path.name}')

def main():
    runtime = llmj.create_bedrock_runtime()
    for path in sorted(llmj.DIR_WORK.glob('*' + llmj.SUFFIX_PRO)):
        process_prompt(runtime, path)
    llmj.finalize()

if __name__ == '__main__':
    main()

