#!/usr/bin/env python3

import sys
import pathlib
import importlib.util

sys.dont_write_bytecode = True
import llmj

try:
    import openpyxl
except ImportError:
    llmj.abort_missing_package('openpyxl')

ARGS = llmj.configure({
    'source' : {
        'default' : str(llmj.DIR_SOURCE),
        'convert' : lambda in_src: pathlib.Path(in_src),
        'explain' : 'Directory containing source text files.'
    },
    'work' : {
        'default' : str(llmj.DIR_WORK),
        'convert' : lambda in_src: pathlib.Path(in_src),
        'explain' : 'Directory containing prompt templates and generated XLSX files.'
    },
    'postprocRetry' : {
        'default' : '10',
        'convert' : lambda in_cnt: int(in_cnt),
        'explain' : 'Maximum number of retries when postproc() requests regeneration.'
    }
})

def load_postproc(in_prompt_path):
    path = in_prompt_path.with_suffix('.postproc.py')
    if not path.exists():
        return lambda in_text: in_text
    spec = importlib.util.spec_from_file_location(path.stem, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    if not hasattr(module, 'postproc'):
        llmj.abort(f'ERROR : "postproc" is not defined in "{path.name}"')
    return module.postproc

def process_prompt(in_prompt_path):
    xls = in_prompt_path.name.removesuffix(llmj.SUFFIX_TXT) + llmj.SUFFIX_XLS
    xls_path = in_prompt_path.with_name(xls)
    if xls_path.exists():
        workbook = openpyxl.load_workbook(xls_path)
    else:
        workbook = openpyxl.Workbook()
    sheet = workbook.active
    colDict = {}
    for key in llmj.TERM_GEN:
        colDict[key] = llmj.find_append_column(sheet, llmj.TERM_GEN[key])
    sourceArr = sorted(ARGS['source'].glob('*.txt'))
    postproc = load_postproc(in_prompt_path)
    for row, src_path in enumerate(sourceArr, start=2):
        print(f'INFO : processing  {row - 1} / {len(sourceArr)}')
        if sheet.cell(row, colDict['GENERATED']).value:
            continue
        textDict = {}
        try:
            with open(src_path, encoding='utf-8') as f:
                textDict['ORIGINAL'] = f.read()
        except Exception:
            print(f'ERROR : can not read "{src_path.name}"')
            llmj.abort()
        prompt = llmj.text_from_template_path(in_prompt_path, {llmj.ORIGINAL_PLACEHOLDER : textDict['ORIGINAL']})
        temperature = None
        for retry in range(ARGS['postprocRetry']):
            generated = postproc(llmj.RUNNER.toText(prompt, None, temperature))
            if generated is not None:
                textDict['GENERATED'] = generated
                break
            temperature = 0.8
            print(f'WARN : retrying because postproc returned None ({retry + 1}/{ARGS["postprocRetry"]})')
        else:
            llmj.abort(f'ERROR : postproc failed after {ARGS["postprocRetry"]} retries')
        for key in llmj.TERM_GEN:
            sheet.cell(row, colDict[key]).value = textDict[key]
        workbook.save(xls_path)
    print(f'INFO : generated {xls_path.name}')

def main():
    for path in sorted(ARGS['work'].glob('*' + llmj.SUFFIX_TXT)):
        process_prompt(path)
    llmj.finalize()

if __name__ == '__main__':
    main()

