#!/usr/bin/env python3

import sys
import json

import llmj

try:
    import openpyxl
except ImportError:
    llmj.abort_missing_package('openpyxl')

REPORT_HTML = r'''
<html>
<head>
<meta charset='utf-8'>
<title>LLMJ Report</title>
<style>

body {
    font-family: sans-serif;
    margin: 20px;
}

table {
    border-collapse: collapse;
    margin-bottom: 24px;
}

th, td {
    border: solid 1px #ccc;
    padding: 6px 10px;
    text-align: left;
    vertical-align: top;
}

#summaryTable td,
#summaryTable th,
#rubricAvgTable td,
#rubricAvgTable th,
#rubricMinTable td,
#rubricMinTable th {
    width: 100px;
}

th.sortable {
    cursor: pointer;
}

pre {
    margin: 0;
    white-space: pre-wrap;
    word-break: break-word;
}

.version {
    margin-bottom: 48px;
}

.article {
    border: solid 1px #ccc;
    padding: 12px;
    margin-bottom: 16px;
}

.reason-link {
    margin-left: 4px;
    white-space: nowrap;
}

.reason-link svg {
    vertical-align: middle;
}

</style>
</head>
<body>
<h1>LLMJ Report</h1>
<div id='report'></div>
<template  id='translateSvg'>
	<svg xmlns='http://www.w3.org/2000/svg' width='14' height='14' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'>
		<path d='m5 8 6 6' />
		<path d='m4 14 6-6 2-3' />
		<path d='M2 5h12' />
		<path d='M7 2h1' />
		<path d='m22 22-5-10-5 10' />
		<path d='M14 18h6' />
	</svg>
</template>
<script>

const gJudgedArr = __JSON__;
const gRubricNameArr = gJudgedArr[0].articleArr[0].results.map(in_r => in_r.name);

/*

[
	{
		name : 'xxx'
		totalAvg : 9.9
		articleArr : [
			{
				original : 'xxx',
				generated : 'xxx',
				average : 9.9,
				results : [
					{
						name : 'xxx',
						score : 9.9,
						reason : 'xxx'
					},
					...
				]
			},
			...
		]
	},
	...
]

*/

function escapeHtml(in_text) {
	const replaceArr = [
		{
			src : '&',
			dst : '&amp;'
		},
		{
			src : '<',
			dst : '&lt;'
		},
		{
			src : '>',
			dst : '&gt;'
		}
	];
	let replaced = in_text;
	replaceArr.forEach(in_replace => {
		replaced = replaced.replaceAll(in_replace.src, in_replace.dst);
	});
	return replaced;
}

function sortTable(in_tblId, in_col) {
	const table = document.getElementById(in_tblId);
	const tbody = table.querySelector('tbody');
	const rowArr = Array.from(tbody.querySelectorAll('tr'));
	const isAsc = table.dataset.sortCol != in_col || table.dataset.sortDir !== 'ascending';
	const extract = in_text => Number(in_text.match(/\d+/)?.[0] ?? 0);
	rowArr.sort((a, b) => {
		const diff = extract(a.children[in_col].textContent) - extract(b.children[in_col].textContent);
		return isAsc ? diff : diff * -1;
	});
	rowArr.forEach(row => tbody.appendChild(row));
	table.dataset.sortCol = in_col;
	table.dataset.sortDir = isAsc ? 'ascending' : 'descending';
}

function sortableTH(in_tblId, in_col, in_title) {
	return '<th class="sortable" data-table="' + in_tblId + '" data-col="' + in_col + '">' + escapeHtml(in_title) + '</th>';
}

function buildRubricTable(in_tblId, in_title, in_callback) {
	let fragment = '';
	fragment += '<h2>' + escapeHtml(in_title) + '</h2>';
	fragment += '<table id="' + in_tblId + '">';
	fragment += '<thead>';
	fragment += '<tr>';
	fragment += sortableTH(in_tblId, 0, 'Version');
	let rubricCol = 1;
	gRubricNameArr.forEach(in_name => fragment += sortableTH(in_tblId, rubricCol++, in_name));
	fragment += '</tr>';
	fragment += '</thead>';
	fragment += '<tbody>';
	gJudgedArr.forEach(in_ver => {
		fragment += '<tr>';
		fragment += '<td>' + escapeHtml(in_ver.name) + '</td>';
		gRubricNameArr.forEach(in_name => {
			const scoreArr = [];
			const min = {
				score : Infinity,
				index : -1
			};
			in_ver.articleArr.forEach((in_art, in_ix) => {
				in_art.results.forEach(in_res => {
					if (in_res.name !== in_name) {
						return;
					}
					const score = Number(in_res.score);
					scoreArr.push(score);
					if (score < min.score) {
						min.score = score;
						min.index = in_ix;
					}
				});
			});
			fragment += in_callback({
				version : in_ver,
				scores : scoreArr,
				minScore : min.score,
				minIndex : min.index
			});
		});
		fragment += '</tr>';
	});
	fragment += '</tbody>';
	fragment += '</table>';
	return fragment;
}

(function() {
	let fragment = '';
	fragment += '<h2>Summary</h2>';
	fragment += '<table id="summaryTable">';
	fragment += '<thead>';
	fragment += '<tr>';
	fragment += sortableTH('summaryTable', 0, 'Version');
	fragment += sortableTH('summaryTable', 1, 'Total Avg');
	fragment += sortableTH('summaryTable', 2, 'Articles');
	fragment += '</tr>';
	fragment += '</thead>';
	fragment += '<tbody>';
	gJudgedArr.forEach(in_ver => {
		fragment += '<tr>';
		fragment += '<td>' + escapeHtml(in_ver.name) + '</td>';
		fragment += '<td data-sort="' + in_ver.totalAvg + '">' + in_ver.totalAvg.toFixed(2) + '</td>';
		fragment += '<td data-sort="' + in_ver.articleArr.length + '">' + in_ver.articleArr.length + '</td>';
		fragment += '</tr>';
	});
	fragment += '</tbody>';
	fragment += '</table>';
	const articleCnt = gJudgedArr[0].articleArr.length;
	fragment += buildRubricTable(
		'rubricAvgTable',
		'Rubric Avg for ' + articleCnt + ' Articles',
		({scores}) => {
			const avg = scores.reduce((a, b) => a + b, 0) / scores.length;
			return '<td data-sort="' + avg + '">' + avg.toFixed(2) + '</td>';
		}
	);
	fragment += buildRubricTable(
		'rubricMinTable',
	    'Rubric Min for ' + articleCnt + ' Articles',
		({version, minScore, minIndex}) => {
			return '<td data-sort="' + minScore + '">' +
					'<a href="#article-' + version.name + '-' + minIndex + '">' + minScore.toFixed(2) + '</a></td>';
		}
	);
	fragment += '<h2>Details</h2>';
	gJudgedArr.forEach(in_ver => {
		fragment += '<div class="version">';
		fragment += `<h3>${escapeHtml(in_ver.name)}</h3>`;
		in_ver.articleArr.forEach((in_art, in_ix) => {
			fragment += '<div class="article" id="article-' + in_ver.name + '-' + in_ix + '">';
			fragment += '<h4>#' + (in_ix + 1) + ' : average=' + in_art.average + '</h4>';
			fragment += '<h5>Original</h5>';
			fragment += '<pre>' + escapeHtml(in_art.original || '') + '</pre>';
			fragment += '<h5>Generated</h5>';
			fragment += '<pre>' + escapeHtml(in_art.generated || '') + '</pre>';
			fragment += '<h5>Rubrics</h5>';
			fragment += '<table>';
			fragment += '<tr>';
			fragment += '<th>Rubric</th>';
			fragment += '<th>Score</th>';
			fragment += '<th>Reason</th>';
			fragment += '</tr>';
			in_art.results.forEach(in_res => {
				const translateUrl =
					'https://translate.google.com/?sl=en&tl=ja&text=' +
					encodeURIComponent(in_res.reason || '') +
					'&op=translate';
				fragment += '<tr>';
				fragment += '<td>' + escapeHtml(in_res.name);
				fragment += '<a class="reason-link" href="' + translateUrl + '" target="_blank" title="Translate">';
				fragment += document.getElementById('translateSvg').innerHTML;
				fragment += '</a>';
				fragment += '</td>';
				fragment += '<td>' + in_res.score + '</td>';
				fragment += '<td>';
				fragment += '<pre>' + escapeHtml(in_res.reason) + '</pre>';
				fragment += '</td>';
				fragment += '</tr>';
			});
			fragment += '</table>';
			fragment += '</div>';
		});
		fragment += '</div>';
	});
	document.getElementById('report').innerHTML = fragment;
	document.querySelectorAll('th.sortable').forEach(th => {
		th.addEventListener('click', () => sortTable(th.dataset.table, Number(th.dataset.col)));
	});
})();

</script>
</body>
</html>
'''

def build_judged_dataset(in_path):
    workbook = openpyxl.load_workbook(in_path, data_only=True)
    sheet = workbook.active
    headDict = {}
    for key in llmj.TERM_ALL:
        headDict[key] = llmj.find_column(sheet, llmj.TERM_ALL[key])
    jsRowArr = []
    for row in range(2, sheet.max_row + 1):
        colDict = {}
        for key in llmj.TERM_ALL:
            jsKey = llmj.TERM_ALL[key]
            colDict[jsKey] = sheet.cell(row, headDict[key]).value
        jsKey = llmj.TERM_ALL['RESULTS']
        colDict[jsKey] = json.loads(colDict[jsKey] or '[]')
        jsRowArr.append(colDict)
    return {
        'name' : in_path.name.removesuffix(llmj.SUFFIX_JUD),
        'totalAvg' : sum(row['average'] for row in jsRowArr) / len(jsRowArr),
        'articleArr' : jsRowArr
    }

def build_html(in_judgedArr):
    return REPORT_HTML.replace('__JSON__', json.dumps(in_judgedArr, ensure_ascii=False))

def main():
    judgedArr = []
    for path in sorted(llmj.DIR_WORK.glob('*' + llmj.SUFFIX_JUD)):
        judgedArr.append(build_judged_dataset(path))
    if len(judgedArr) > 0:
        out_path = llmj.DIR_WORK / llmj.FILE_REPORT
        print(f'processing : {out_path.name}')
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(build_html(judgedArr))
    llmj.finalize()

if __name__ == '__main__':
    main()
