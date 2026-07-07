#!/usr/bin/env python3

import os
import sys
import time
import json
import dotenv
import shutil
import pathlib
import concurrent.futures

def abort(in_message=None):
    if in_message:
        print(in_message)
    finalize()
    sys.exit(1)

dotenv.load_dotenv()
for required in ['ACCESS_KEY_ID', 'SECRET_ACCESS_KEY', 'SESSION_TOKEN', 'GATEWAY_URL']:
    if os.getenv(required) is None:
        print(f'ERROR : {required} is not defined in ".env".')
        abort()

def abort_missing_package(in_package):
    print(f'ERROR : exec "python -m pip install {in_package}" at first.')
    abort()

try:
    import boto3
except ImportError:
    abort_missing_package('boto3')

try:
    import openpyxl
except ImportError:
    abort_missing_package('openpyxl')

try:
    import langdetect
except ImportError:
    abort_missing_package('langdetect')

try:
    from deepeval.metrics import GEval
    from deepeval.test_case import LLMTestCase
    from deepeval.test_case import SingleTurnParams
    from deepeval.models import DeepEvalBaseLLM
    os.environ['DEEPEVAL_TELEMETRY_OPT_OUT'] = 'YES'
except ImportError:
    abort_missing_package('deepeval')

DIR_ROOT = pathlib.Path(__file__).resolve().parent
DIR_WORK = DIR_ROOT / 'work'
DIR_SOURCE = DIR_ROOT / 'source'
DIR_RUBRIC = DIR_ROOT / 'rubric'
DIR_SUPPORTS = DIR_ROOT / 'supports'

for path in [DIR_SOURCE, DIR_RUBRIC, DIR_SUPPORTS]:
    if not path.is_dir():
        print(f'ERROR : can not find "{path.name}" directory.')
        abort()

DIR_WORK.mkdir(exist_ok=True)

_QUOTATION = {
    'ASCII' : chr(0x22),
    'FULLW' : chr(0xFF02)
}

_APOSTROPHE = {
    'ASCII' : chr(0x27),
    'FULLW' : chr(0xFF07)
}

TERM_GEN = {
    'ORIGINAL' : 'original',
    'GENERATED' : 'generated',
}

TERM_JUD = {
    'AVERAGE' : 'average',
    'RESULTS' : 'results'
}

TERM_ALL = TERM_GEN | TERM_JUD

# prompt
SUFFIX_TXT = '.txt'
# generated / judged
SUFFIX_XLS = '.xlsx'

INITIAL_VERSION_NAME = 'initial-prompt'

ORIGINAL_PLACEHOLDER = '{{' + TERM_ALL['ORIGINAL'] + '}}'

def _create_finalize():
    start_time = time.time()
    def _finalize():
        for name in ['.deepeval']:
            shutil.rmtree(pathlib.Path.cwd() / name, ignore_errors=True)
        elapsed = time.time() - start_time
        print(f'INFO : completed {pathlib.Path(sys.argv[0]).name} ( elapsed : {elapsed:.1f} sec )')
    return _finalize

finalize = _create_finalize()

def _find_column(in_sheet, in_name):
    for col in range(1, in_sheet.max_column + 1):
        value = in_sheet.cell(row=1, column=col).value
        if value == in_name:
            return col
    return None

def find_append_column(in_sheet, in_name):
    col = _find_column(in_sheet, in_name)
    if col is not None:
        return col
    col = in_sheet.max_column + 1
    in_sheet.cell(row=1, column=col).value = in_name
    return col

class cLLMRunner:
    def __init__(
        self,
        in_model='us.anthropic.claude-sonnet-4-6',
        in_region='us-east-1',
        in_maxTokens=4096,
        in_temperature=0,
        in_retryCount=3,
        in_retryInterval=5
    ):
        self._model = in_model
        self._region = in_region
        self._maxTokens = in_maxTokens
        self._temperature = in_temperature
        self._retryCount = in_retryCount
        self._retryInterval = in_retryInterval
        self._runtime = self._create_runtime()
    @property
    def model(self):
        return self._model
    def _create_runtime(self):
        dotenv.load_dotenv()
        for required in [
            'ACCESS_KEY_ID',
            'SECRET_ACCESS_KEY',
            'SESSION_TOKEN',
            'GATEWAY_URL'
        ]:
            if os.getenv(required) is None:
                abort(f'ERROR : {required} is not defined in ".env".')
        session = boto3.Session(
            aws_access_key_id=os.getenv('ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('SECRET_ACCESS_KEY'),
            aws_session_token=os.getenv('SESSION_TOKEN')
        )
        return session.client(
            'bedrock-runtime',
            region_name=self._region,
            endpoint_url=os.getenv('GATEWAY_URL')
        )
    def _retry(self, in_callback):
        for retry in range(self._retryCount):
            try:
                return in_callback()
            except Exception as err:
                if retry + 1 >= self._retryCount:
                    abort(f'ERROR : invoke failed : {err}')
                print(f'WARN : retrying because : {err}')
                time.sleep(self._retryInterval * (retry + 1))
    def _invoke(self, in_prompt, in_maxTokens, in_temperature):
        if in_maxTokens is None:
            in_maxTokens = self._maxTokens
        if in_temperature is None:
            in_temperature = self._temperature
        def callback():
            response = self._runtime.converse_stream(
                modelId=self._model,
                messages=[{
                    'role' : 'user',
                    'content' : [{'text' : in_prompt}]
                }],
                inferenceConfig={'maxTokens' : in_maxTokens, 'temperature' : in_temperature}
            )
            chunkArr = []
            for event in response['stream']:
                if 'contentBlockDelta' not in event:
                    continue
                delta = event['contentBlockDelta']['delta']
                if 'text' not in delta:
                    continue
                chunkArr.append(delta['text'])
            return ''.join(chunkArr)
        return self._retry(callback)
    def toText(self,
        in_prompt,
        in_maxTokens=None,
        in_temperature=None
    ):
        return self._invoke(in_prompt, in_maxTokens, in_temperature)
    def toJson(self,
        in_prompt,
        in_maxTokens=None,
        in_temperature=None
    ):
        try:
            return json.loads(self._invoke(in_prompt, in_maxTokens, in_temperature))
        except Exception as err:
            abort(f'ERROR : invalid json : {err}')

RUNNER = cLLMRunner()

def text_from_template_text(in_text, in_replaceDict):
    text = in_text
    for placeholder, replaced in in_replaceDict.items():
        if isinstance(replaced, (dict, list)):
            replaced = json.dumps(replaced, ensure_ascii=False, indent=2)
        else:
            replaced = str(replaced)
        text = text.replace(placeholder, replaced)
    return text

def text_from_template_path(in_path, in_replaceDict):
    with open(in_path, encoding='utf-8') as f:
        text = f.read()
    return text_from_template_text(text, in_replaceDict)

def llm_processed_text(in_path, in_replaceDict):
    return RUNNER.toText(text_from_template_path(in_path, in_replaceDict))

def llm_processed_json(in_path, in_replaceDict):
    return RUNNER.toJson(text_from_template_path(in_path, in_replaceDict))

def load_rubrics():
    rubricArr = []
    for path in sorted(DIR_RUBRIC.glob('*.json')):
        try:
            with open(path, encoding='utf-8') as f:
                rubricArr.append(json.load(f))
        except Exception:
            return None
    if len(rubricArr) == 0:
        return None
    return rubricArr

class _GatewayLLM(DeepEvalBaseLLM):
    def __init__(self, in_runner):
        self.runner = in_runner
    def get_model_name(self):
        return self.runner.model
    def load_model(self):
        return self
    def generate(self, in_prompt):
        return self.runner.toText(in_prompt)
    async def a_generate(self, in_prompt):
        return self.generate(in_prompt)

#invoke_llm

def _create_judge(in_rubricArr):
    def evaluate(in_rubric, in_testcase):
        metric = GEval(
            name=in_rubric['name'],
            # criteria=in_rubric['criteria'],
            evaluation_steps=in_rubric['evaluation_steps'],
            evaluation_params=[
                SingleTurnParams.INPUT,
                SingleTurnParams.ACTUAL_OUTPUT
            ],
            async_mode=False,
            model=_GatewayLLM(RUNNER)
        )
        metric.measure(in_testcase)
        resDict = {}
        for key in ['name', 'score', 'reason']:
            resDict[key] = getattr(metric, key)
        return resDict
    def judge(in_original, in_generated):
        testcase = LLMTestCase(
            input=in_original,
            actual_output=in_generated
        )
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(in_rubricArr)) as executor:
            callback = lambda in_rubric: evaluate(in_rubric, testcase)
            resultArr = list(executor.map(callback, in_rubricArr))
        scoreArr = []
        for result in resultArr:
            scoreArr.append(result['score'])
        return {
            TERM_ALL['AVERAGE'] : sum(scoreArr) / len(scoreArr),
            TERM_ALL['RESULTS'] : json.dumps(resultArr, ensure_ascii=False)
        }
    return judge

def _is_judged_row(in_sheet, in_row, in_colDict):
    for key in TERM_JUD:
        if in_sheet.cell(in_row, in_colDict[key]).value is None:
            return False
    return True

def _is_judged_xlsx(in_path):
    workbook = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
    sheet = workbook.active
    colDict = {}
    for key in TERM_JUD:
        col = _find_column(sheet, TERM_ALL[key])
        if col is None:
            # judge columns do not exist yet ( before the first judge )
            return False
        colDict[key] = col
    for row in range(2, sheet.max_row + 1):
        if not _is_judged_row(sheet, row, colDict):
            # some rows are still unjudged
            return False
    return True

def _process_xlsx(in_path, in_callback):
    workbook = openpyxl.load_workbook(in_path)
    sheet = workbook.active
    colDict = {}
    for key in TERM_ALL:
        colDict[key] = find_append_column(sheet, TERM_ALL[key])
    for row in range(2, sheet.max_row + 1):
        print(f'INFO : processing {row - 1} / {sheet.max_row - 1}')
        if _is_judged_row(sheet, row, colDict):
            continue
        # replace characters that DeepEval cannot handle
        safeText = {}
        for key in TERM_GEN:
            safeText[key] = sheet.cell(row, colDict[key]).value
            for repDict in [_QUOTATION, _APOSTROPHE]:
                safeText[key] = safeText[key].replace(repDict['ASCII'], repDict['FULLW'])
        result = in_callback(safeText['ORIGINAL'], safeText['GENERATED'])
        for key in TERM_JUD:
            sheet.cell(row, colDict[key]).value = result[TERM_ALL[key]]
        workbook.save(in_path)
    print(f'INFO : judged {in_path.name}')

def _build_judged_dataset(in_path):
    workbook = openpyxl.load_workbook(in_path, data_only=True)
    sheet = workbook.active
    headDict = {}
    for key in TERM_ALL:
        headDict[key] = _find_column(sheet, TERM_ALL[key])
    jsRowArr = []
    for row in range(2, sheet.max_row + 1):
        colDict = {}
        colDict['articleIndex'] = row - 2
        for key in TERM_ALL:
            jsKey = TERM_ALL[key]
            colDict[jsKey] = sheet.cell(row, headDict[key]).value
        try:
            colDict['lang'] = langdetect.detect(colDict[TERM_ALL['ORIGINAL']])
        except Exception:
            colDict['lang'] = 'en'
        jsKey = TERM_ALL['RESULTS']
        colDict[jsKey] = json.loads(colDict[jsKey] or '[]')
        jsRowArr.append(colDict)
    return {
        'name' : in_path.name.removesuffix(SUFFIX_XLS),
        'totalAvg' : sum(row['average'] for row in jsRowArr) / len(jsRowArr),
        'articleArr' : jsRowArr
    }

def build_judged_dataset_array():
    rubricArr = load_rubrics()
    if rubricArr is None:
        print('ERROR : can not read some json')
        abort()
    judgeCallback = None
    judgedArr = []
    for path in sorted(DIR_WORK.glob('*' + SUFFIX_XLS)):
        if not _is_judged_xlsx(path):
            if judgeCallback is None:
                print(f'INFO : compiling {len(rubricArr)} rubrics')
                compiledArr = llm_processed_json(DIR_SUPPORTS / 'template-compiler.txt', {'__JSON__' : rubricArr})
                judgeCallback = _create_judge(compiledArr)
            _process_xlsx(path, judgeCallback)
        judgedArr.append(_build_judged_dataset(path))
    return judgedArr
