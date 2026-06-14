#!/usr/bin/env python3

import os
import io
import sys
import json
import time
import contextlib
import concurrent.futures

import llmj

try:
    import openpyxl
except ImportError:
    llmj.abort_missing_package('openpyxl')

try:
    from deepeval.metrics import GEval
    from deepeval.test_case import LLMTestCase
    from deepeval.test_case import SingleTurnParams
    from deepeval.models import DeepEvalBaseLLM
    os.environ['DEEPEVAL_TELEMETRY_OPT_OUT'] = 'YES'
except ImportError:
    llmj.abort_missing_package('deepeval')

def load_rubrics():
    rubricArr = []
    for path in sorted(llmj.DIR_RUBRIC.glob('*.json')):
        try:
            with open(path, encoding='utf-8') as f:
                rubricArr.append(json.load(f))
        except Exception:
            return None
    if len(rubricArr) == 0:
        return None
    return rubricArr

def compile_rubric(in_rubricArr):
    promptToCompile = f'''

You are an expert evaluator designer for DeepEval GEval.
Convert the criteria into evaluation_steps optimized for GEval.

[Requirements]

- Produce explicit evaluation procedures, not evaluation rules.
- Write steps in the order they should be executed.
- Avoid redundant or overlapping checks.
- Prefer fact extraction --> comparison --> scoring workflow.
- Include all important constraints from the criteria.
- The steps must be reusable across many test cases.
- Return only JSON.
  - Preserve all existing fields.
  - Add an evaluation_steps field to each rubric.
- Do not use markdown.
- Do not wrap the response in code fences.

[Expected]

[
    {{
        "name": "...",
        "criteria": "...",
        "evaluation_steps": [...]
    }},
    {{
        "name": "...",
        "criteria": "...",
        "evaluation_steps": [...]
    }},
    ...
]

[Input]

{json.dumps(in_rubricArr, ensure_ascii=False, indent=2)}

'''.strip()

    runtime = llmj.create_bedrock_runtime()
    generated = llmj.invoke_llm(runtime, llmj.LLM_MODEL, promptToCompile)
    try:
        return json.loads(generated)
    except Exception:
        print('ERROR : failed to compile rubric')
        llmj.abort()

class GatewayLLM(DeepEvalBaseLLM):
    def __init__(self, in_runtime, in_model):
        self.runtime = in_runtime
        self.model = in_model
    def get_model_name(self):
        return self.model
    def load_model(self):
        return self
    def generate(self, in_prompt):
        return llmj.invoke_llm(self.runtime, self.model, in_prompt)
    async def a_generate(self, in_prompt):
        return self.generate(in_prompt)

def create_judge(in_rubricArr):
    runtime = llmj.create_bedrock_runtime()
    def evaluate(in_rubric, in_testcase):
        metric = GEval(
            name=in_rubric['name'],
            # criteria=in_rubric['criteria'],
            evaluation_steps=in_rubric['evaluation_steps'],
            evaluation_params=[
                SingleTurnParams.INPUT,
                SingleTurnParams.ACTUAL_OUTPUT
            ],
            async_mode=False,
            model=GatewayLLM(runtime, llmj.LLM_MODEL)
        )
        llmj.invoke(lambda: metric.measure(in_testcase))
        resDict = {}
        for key in ['name', 'score', 'reason']:
            resDict[key] = getattr(metric, key)
        return resDict
    def judge(in_original, in_generated):
        testcase = LLMTestCase(
            input=in_original,
            actual_output=in_generated
        )
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(in_rubricArr)) as executor:
            callback = lambda in_rubric: evaluate(in_rubric, testcase)
            resultArr = list(executor.map(callback, in_rubricArr))
        scoreArr = []
        for result in resultArr:
            scoreArr.append(result['score'])
        return {
            llmj.TERM_ALL['AVERAGE'] : sum(scoreArr) / len(scoreArr),
            llmj.TERM_ALL['RESULTS'] : json.dumps(resultArr, ensure_ascii=False)
        }
    return judge

def process_xlsx(in_path, in_callback):
    out_path = in_path.with_name(in_path.name.removesuffix(llmj.SUFFIX_GEN) + llmj.SUFFIX_JUD)
    if not out_path.exists():
        workbook = openpyxl.load_workbook(in_path)
        workbook.save(out_path)
    workbook = openpyxl.load_workbook(out_path)
    sheet = workbook.active
    colDict = {}
    for key in llmj.TERM_ALL:
        colDict[key] = llmj.find_append_column(sheet, llmj.TERM_ALL[key])
    for row in range(2, sheet.max_row + 1):
        print(f'processing : {row - 1} / {sheet.max_row - 1}')
        skip = True
        for key in llmj.TERM_JUD:
            skip = skip and sheet.cell(row, colDict[key]).value is not None
        if skip:
            continue
        # replace characters that DeepEval cannot handle
        safeText = {}
        for key in llmj.TERM_GEN:
            safeText[key] = sheet.cell(row, colDict[key]).value
            for repDict in [llmj.QUOTATION, llmj.APOSTROPHE]:
                safeText[key] = safeText[key].replace(repDict['ASCII'], repDict['FULLW'])
        result = in_callback(safeText['ORIGINAL'], safeText['GENERATED'])
        for key in llmj.TERM_JUD:
            sheet.cell(row, colDict[key]).value = result[llmj.TERM_ALL[key]]
        workbook.save(out_path)
    print(f'judged : {out_path.name}')

def main():
    rubricArr = load_rubrics()
    if rubricArr is None:
        print('ERROR : can not read some json')
        llmj.abort()
    callback = create_judge(compile_rubric(rubricArr))
    for path in sorted(llmj.DIR_WORK.glob('*' + llmj.SUFFIX_GEN)):
        process_xlsx(path, callback)
    llmj.finalize()

if __name__ == '__main__':
    main()
